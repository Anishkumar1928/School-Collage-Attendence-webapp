from datetime import datetime
import os
import openpyxl
from datetime import datetime
import itertools
import shutil
#extract student from excel sheet
def extrat_stu(file,section):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file)
    sheet = wb[f'{section}']
    totalstudent=[]    
    for row in range(2, sheet.max_row):
        if sheet.cell(row=row, column=1).value==None:
            continue
        else:
            totalstudent.append(sheet.cell(row=row, column=1).value)
    return totalstudent
#take attendence logic function which use to write on excel sheet 
#whose arguments are excelfilename,listof present student list of absent student
def take_attendance(filepath,section,pr,ab):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[f'{section}']

    # Get today's date
    today = datetime.now().date()
    today_column =None
    #search new cloumn for new day aatendence and assign first cell today date
    #and vvvvv here colum is today,s column
    for colum in itertools.count(start=2):
        if sheet.cell(row=1, column=colum).value==str(today):
            today_column=colum
            break
        elif sheet.cell(row=1, column=colum).value==None:
            sheet.cell(row=1,column=colum,value=str(today))
            today_column=colum
            break
    #this for chacking our new coloumn has value today,s date or not
    """for c in range(2, sheet.max_column +1):
        print(sheet.cell(row=1, column=c).value)"""   


    # Find the column corresponding to today's date
    for row in range(2, sheet.max_row + 1):
        #if student name is present in absent list
        if sheet.cell(row=row, column=1).value in ab:
            #write "A" in today coloumn and  name row   
            sheet.cell(row=row, column=today_column, value='A')
        elif sheet.cell(row=row, column=1).value in pr:
            sheet.cell(row=row, column=today_column, value='P')
    #here we save the excel file and print a message Attendance saved :         
    wb.save(filepath)
    print("Attendance saved.")

def editatt(filepath,section,name,date,cellvalue,user):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[f'{section}']
    datecolum=None
    #find date column
    for colum in range(2, sheet.max_column+1):
        if sheet.cell(row=1, column=colum).value==date:
            datecolum=colum
            break
    if datecolum==None:
        return "date not found"
    else:
        print(datecolum)   
        for row in range(2, sheet.max_row + 1):
            #if student name is present in absent list
            if sheet.cell(row=row, column=1).value == name:
                sheet.cell(row=row, column=datecolum).value=cellvalue        
        wb.save(filepath)
        with open('editattendence_log.txt','a') as file:
            file.write(f"\n{datetime.now()} {user} edit the attendence of {name} of date{date}")
        return f"{name} attendence updated acivity is saved"

def copyfile():
    ogf="3SEMBCA.xlsx"
    cpyf="cpoy.xlsx"
    shutil.copy(ogf, cpyf)
            # Send the copy of the Excel file to the user
    with open(cpyf, 'rb') as file:
        return file
    
def copysheets(path,pathsheet,sheetname,savepath):
    wb=openpyxl.load_workbook(path)
    sheet=wb[pathsheet]
    dwb=openpyxl.Workbook()

    sheetname=dwb.create_sheet(title=sheetname)
    #dwb.remove_sheet('Sheet')
    if 'Sheet' in dwb.sheetnames:
        rmvsh=dwb['Sheet'] 
        dwb.remove(rmvsh)
    os.remove(savepath)

    for row in sheet.iter_rows(values_only=True):
        sheetname.append(row)
    dwb.save(savepath)



def nosheets(file):
    workbook=openpyxl.load_workbook(file)
    sheet_names=workbook.sheetnames
    return sheet_names


# print(editatt("AWEB_DEV","Ayushi Shrivastav",'2023-12-31',"sro")) 
"""if __name__ == "__main__":""""""pr=[]
ab=["KRITI KUMARI","HIMANSHU KUMAR","HARSH RAJ","AAKASH RAJ","SAUMYA KUMARI","ANSHU KUMARI","PANKAJ KUMAR SAH"]
    file_path = "attendance.xlsx"""
    #take_attendance(file_path,pr,ab)"""