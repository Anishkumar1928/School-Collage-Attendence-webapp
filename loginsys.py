import openpyxl
from openpyxl import workbook


def loginu(userid,password):
    wb = openpyxl.load_workbook('data base.xlsx')
    sheet = wb['Sheet1']
    id_=[]    
    for row in range(2,sheet.max_row+1):
        user_value=str(sheet.cell(row=row, column=1).value)
        pass_value=str(sheet.cell(row=row, column=2).value)
        name_value=str(sheet.cell(row=row, column=4).value)
        if user_value==userid and password==pass_value:
            return True,name_value
