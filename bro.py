import os
import openpyxl
from datetime import datetime
import itertools
import shutil

def copysheets(path,pathsheet,name,sheetname):
    wb=openpyxl.load_workbook(path)
    sheet=wb[pathsheet]
    dwb=openpyxl.Workbook()

    sheetname=dwb.create_sheet(title=sheetname)
    #dwb.remove_sheet('Sheet')
    if 'Sheet' in dwb.sheetnames:
        rmvsh=dwb['Sheet'] 
        dwb.remove(rmvsh)

    for row in sheet.iter_rows(values_only=True):
        sheetname.append(row)
    dwb.save(name)

copysheets('data base.xlsx','Sheet1','anish.xlsx','A')

def copyfile(path):
    ogf=path
    cpyf="./static/copy.xlsx"
    shutil.copy(ogf, cpyf)
            # Send the copy of the Excel file to the user
#     with open(cpyf, 'rb') as file:
#         return file


# bot.send_document(message.chat.id, file, caption="Attendance Report")
        # Delete the copy file (optional)
# os.remove(copy_file_path)

